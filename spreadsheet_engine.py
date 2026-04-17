
from __future__ import annotations
from dataclasses import dataclass
from datetime import date, datetime
from calendar import monthrange
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re

from openpyxl import load_workbook


CATEGORY_SYNONYMS = {
    "casa": "Moradia",
    "moradia": "Moradia",
    "carro": "Carro",
    "internet": "Internet e Celular",
    "celular": "Internet e Celular",
    "internet e celular": "Internet e Celular",
    "telefone": "Internet e Celular",
    "luz": "Luz e Streaming",
    "streaming": "Luz e Streaming",
    "energia": "Luz e Streaming",
    "luz e streaming": "Luz e Streaming",
    "pet": "Pet",
    "ração": "Pet",
    "racao": "Pet",
    "cachorra": "Pet",
    "gasolina": "Combustível",
    "combustivel": "Combustível",
    "combustível": "Combustível",
    "lazer": "Lazer",
    "saude": "Saúde e Farmácia",
    "saúde": "Saúde e Farmácia",
    "farmacia": "Saúde e Farmácia",
    "farmácia": "Saúde e Farmácia",
    "imprevisto": "Imprevistos",
    "imprevistos": "Imprevistos",
    "reserva": "Reserva de Emergência",
    "emergencia": "Reserva de Emergência",
    "emergência": "Reserva de Emergência",
    "extra": "Imprevistos",
    "extras": "Imprevistos",
}

@dataclass
class LaunchResult:
    category: str
    amount: float
    month_key: str
    spent: float
    limit: float
    remaining: float
    status: str
    message: str


class FinanceWorkbook:
    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook não encontrado: {self.workbook_path}")

    def _wb(self):
        return load_workbook(self.workbook_path)

    def normalize_category(self, text: str) -> str:
        key = (text or "").strip().lower()
        if key in CATEGORY_SYNONYMS:
            return CATEGORY_SYNONYMS[key]
        for alias, canonical in CATEGORY_SYNONYMS.items():
            if key and (key in alias or alias in key):
                return canonical
        return (text or "").strip().title()

    def format_currency_br(self, value: float) -> str:
        return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def config_map(self) -> Dict[str, str]:
        wb = self._wb()
        ws = wb["Config"]
        data = {}
        row = 4
        while ws[f"A{row}"].value:
            data[str(ws[f"A{row}"].value)] = ws[f"B{row}"].value
            row += 1
        return data

    def alert_days(self) -> int:
        value = self.config_map().get("dias_alerta_contas", 5)
        try:
            return int(value)
        except Exception:
            return 5

    def load_limits(self) -> Dict[str, float]:
        wb = self._wb()
        ws = wb["Limites"]
        limits = {}
        row = 4
        while ws[f"A{row}"].value:
            category = str(ws[f"A{row}"].value).strip()
            if category.lower().startswith("total"):
                break
            limits[category] = float(ws[f"B{row}"].value or 0)
            row += 1
        return limits

    def load_fixed_bills(self) -> List[dict]:
        wb = self._wb()
        ws = wb["Contas_Fixas"]
        bills = []
        row = 4
        while ws[f"A{row}"].value:
            bills.append({
                "name": str(ws[f"A{row}"].value),
                "category": str(ws[f"B{row}"].value),
                "amount": float(ws[f"C{row}"].value or 0),
                "frequency": str(ws[f"D{row}"].value),
                "due_day": int(ws[f"E{row}"].value or 1),
                "bill_type": str(ws[f"F{row}"].value),
                "active": str(ws[f"G{row}"].value).strip().lower() == "sim",
                "note": str(ws[f"H{row}"].value or ""),
            })
            row += 1
        return bills

    def load_launches(self, month_key: Optional[str] = None) -> List[dict]:
        wb = self._wb()
        ws = wb["Lancamentos"]
        launches = []
        row = 4
        while row <= ws.max_row:
            if not ws[f"A{row}"].value:
                row += 1
                continue
            d = ws[f"A{row}"].value
            if isinstance(d, datetime):
                d = d.date()
            if not isinstance(d, date):
                row += 1
                continue
            current_month_key = d.strftime("%Y-%m")
            if month_key and current_month_key != month_key:
                row += 1
                continue
            launches.append({
                "date": d,
                "month": current_month_key,
                "category": str(ws[f"C{row}"].value or "").strip(),
                "subcategory": str(ws[f"D{row}"].value or "").strip(),
                "entry_type": str(ws[f"E{row}"].value or "").strip().lower(),
                "description": str(ws[f"F{row}"].value or "").strip(),
                "value": float(ws[f"G{row}"].value or 0),
                "channel": str(ws[f"H{row}"].value or "").strip(),
                "status": str(ws[f"I{row}"].value or "").strip(),
                "note": str(ws[f"J{row}"].value or "").strip(),
            })
            row += 1
        return launches

    def last_launches(self, limit: int = 8) -> List[dict]:
        launches = self.load_launches()
        launches.sort(key=lambda x: x["date"], reverse=True)
        out = []
        for item in launches[:limit]:
            out.append({
                "date": item["date"].isoformat(),
                "category": item["category"],
                "description": item["description"],
                "value": round(item["value"], 2),
                "entry_type": item["entry_type"],
                "channel": item["channel"],
            })
        return out

    def _find_next_launch_row(self, ws) -> int:
        row = 4
        while row <= max(ws.max_row + 2, 500):
            if not ws[f"A{row}"].value and not ws[f"C{row}"].value and not ws[f"G{row}"].value:
                return row
            row += 1
        return ws.max_row + 1

    def add_launch(
        self,
        category: str,
        amount: float,
        description: str = "",
        entry_type: str = "gasto",
        channel: str = "telegram",
        entry_date: Optional[date] = None,
        status: str = "Pago",
        note: str = "",
        subcategory: str = "",
    ) -> LaunchResult:
        category = self.normalize_category(category)
        entry_date = entry_date or date.today()
        month_key = entry_date.strftime("%Y-%m")
        wb = self._wb()
        ws = wb["Lancamentos"]
        row = self._find_next_launch_row(ws)

        ws[f"A{row}"] = entry_date
        ws[f"B{row}"] = month_key
        ws[f"C{row}"] = category
        ws[f"D{row}"] = subcategory or ("Parcela" if category in {"Moradia", "Carro"} else "")
        ws[f"E{row}"] = entry_type
        ws[f"F{row}"] = description or f"Lançamento via {channel}"
        ws[f"G{row}"] = float(amount)
        ws[f"H{row}"] = channel
        ws[f"I{row}"] = status
        ws[f"J{row}"] = note
        wb.save(self.workbook_path)

        summary = self.category_summary(category, month_key)
        kind_emoji = "💰" if entry_type == "receita" else "✅"
        message = (
            f"{kind_emoji} {category}: lançado {self.format_currency_br(amount)}\n"
            f"Tipo: {entry_type}\n"
            f"Gasto no mês: {self.format_currency_br(summary['spent'])}\n"
            f"Limite: {self.format_currency_br(summary['limit'])}\n"
            f"Saldo: {self.format_currency_br(summary['remaining'])}\n"
            f"Status: {summary['status']}"
        )
        return LaunchResult(
            category=category,
            amount=amount,
            month_key=month_key,
            spent=summary["spent"],
            limit=summary["limit"],
            remaining=summary["remaining"],
            status=summary["status"],
            message=message,
        )

    def category_summary(self, category: str, month_key: Optional[str] = None) -> dict:
        month_key = month_key or date.today().strftime("%Y-%m")
        category = self.normalize_category(category)
        limits = self.load_limits()
        launches = self.load_launches(month_key)
        spent = sum(x["value"] for x in launches if x["category"] == category and x["entry_type"] == "gasto")
        limit = float(limits.get(category, 0))
        remaining = limit - spent
        status = "OK"
        if limit > 0 and spent > limit:
            status = "ESTOURADO"
        elif limit > 0 and spent >= (limit * 0.85):
            status = "ALERTA"
        return {
            "category": category,
            "month": month_key,
            "spent": round(spent, 2),
            "limit": round(limit, 2),
            "remaining": round(remaining, 2),
            "status": status,
        }

    def month_summary(self, month_key: Optional[str] = None) -> dict:
        month_key = month_key or date.today().strftime("%Y-%m")
        limits = self.load_limits()
        launches = self.load_launches(month_key)
        categories = []
        total_limit = 0.0
        total_spent = 0.0
        for category, limit in limits.items():
            spent = sum(x["value"] for x in launches if x["category"] == category and x["entry_type"] == "gasto")
            remaining = limit - spent
            status = "OK"
            if limit > 0 and spent > limit:
                status = "ESTOURADO"
            elif limit > 0 and spent >= limit * 0.85:
                status = "ALERTA"
            categories.append({
                "category": category,
                "limit": round(limit, 2),
                "spent": round(spent, 2),
                "remaining": round(remaining, 2),
                "status": status,
                "used_pct": round((spent / limit) if limit else 0, 4),
            })
            total_limit += limit
            total_spent += spent

        total_remaining = total_limit - total_spent
        receitas = sum(x["value"] for x in launches if x["entry_type"] == "receita")
        overall_status = "OK"
        if total_limit > 0 and total_spent > total_limit:
            overall_status = "ESTOURADO"
        elif total_limit > 0 and total_spent >= total_limit * 0.85:
            overall_status = "ALERTA"
        return {
            "month": month_key,
            "total_limit": round(total_limit, 2),
            "total_spent": round(total_spent, 2),
            "total_remaining": round(total_remaining, 2),
            "used_pct": round((total_spent / total_limit) if total_limit else 0, 4),
            "status": overall_status,
            "extra_income": round(receitas, 2),
            "categories": categories,
        }

    def due_bills(self, days_ahead: int = 5, ref_date: Optional[date] = None) -> List[dict]:
        ref_date = ref_date or date.today()
        bills = []
        for bill in self.load_fixed_bills():
            if not bill["active"]:
                continue
            due_this_month = date(ref_date.year, ref_date.month, min(bill["due_day"], monthrange(ref_date.year, ref_date.month)[1]))
            if bill["frequency"].lower() == "trimestral":
                if ref_date.month not in (1, 4, 7, 10):
                    continue
            if 0 <= (due_this_month - ref_date).days <= days_ahead:
                bills.append({
                    **bill,
                    "due_date": due_this_month.isoformat(),
                    "days_left": (due_this_month - ref_date).days,
                })
        bills.sort(key=lambda x: (x["days_left"], x["due_day"]))
        return bills

    def mark_bill_paid(self, bill_name: str, ref_date: Optional[date] = None) -> LaunchResult:
        ref_date = ref_date or date.today()
        bill_name_lower = bill_name.strip().lower()
        matched = None
        for bill in self.load_fixed_bills():
            if bill_name_lower in bill["name"].lower() or bill_name_lower in bill["category"].lower():
                matched = bill
                break
        if not matched:
            raise ValueError("Conta não encontrada.")
        return self.add_launch(
            category=matched["category"],
            amount=float(matched["amount"]),
            description=matched["name"],
            entry_type="gasto",
            channel="telegram",
            entry_date=ref_date,
            status="Pago",
            note="Conta marcada como paga via comando /pagar",
            subcategory="Parcela" if matched["category"] in {"Moradia", "Carro"} else "",
        )

    def status_message(self, month_key: Optional[str] = None) -> str:
        data = self.month_summary(month_key)
        lines = [
            f"📊 Status {data['month']}",
            f"Total gasto: {self.format_currency_br(data['total_spent'])}",
            f"Limite total: {self.format_currency_br(data['total_limit'])}",
            f"Saldo: {self.format_currency_br(data['total_remaining'])}",
            f"Uso do orçamento: {round(data['used_pct']*100, 1)}%",
            f"Receitas extras: {self.format_currency_br(data['extra_income'])}",
            f"Status geral: {data['status']}",
        ]
        alerts = [c for c in data["categories"] if c["status"] != "OK"]
        if alerts:
            lines.append("")
            lines.append("⚠️ Alertas:")
            for item in alerts[:5]:
                lines.append(f"- {item['category']}: gasto {self.format_currency_br(item['spent'])} / limite {self.format_currency_br(item['limit'])} ({item['status']})")
        return "\n".join(lines)

    def limits_message(self, month_key: Optional[str] = None) -> str:
        data = self.month_summary(month_key)
        lines = [f"📦 Limites {data['month']}"]
        for item in data["categories"]:
            lines.append(f"- {item['category']}: {self.format_currency_br(item['spent'])} / {self.format_currency_br(item['limit'])} | saldo {self.format_currency_br(item['remaining'])}")
        return "\n".join(lines)

    def bills_message(self, days_ahead: int = 5) -> str:
        bills = self.due_bills(days_ahead)
        if not bills:
            return "✅ Nenhuma conta vencendo nos próximos dias."
        lines = ["📅 Contas próximas:"]
        for bill in bills:
            lines.append(f"- {bill['name']} | vence em {bill['due_date']} | {self.format_currency_br(bill['amount'])}")
        return "\n".join(lines)

    def parse_telegram_message(self, text: str) -> Tuple[str, Optional[dict]]:
        raw = (text or "").strip()
        msg = raw.lower()
        if not raw:
            return "Envie um comando. Ex.: /gasto gasolina 120, /status, /contas.", None

        if msg in {"/start", "start"}:
            return (
                "🚀 Financeiro Telegram Premium\n"
                "Use /ajuda para ver os comandos.\n"
                "Exemplo rápido: gasolina 120",
                None,
            )

        if msg in {"/ajuda", "ajuda", "help", "menu"}:
            return (
                "🤖 Comandos disponíveis:\n"
                "/gasto categoria valor descrição\n"
                "/receita categoria valor descrição\n"
                "/status\n"
                "/status categoria\n"
                "/limites\n"
                "/contas\n"
                "/pagar nome da conta\n"
                "/ajuda\n\n"
                "Atalho simples: gasolina 120",
                None,
            )

        if msg in {"/status", "status"}:
            return self.status_message(), None

        if msg.startswith("/status ") or msg.startswith("status "):
            category = self.normalize_category(raw.split(" ", 1)[1])
            data = self.category_summary(category)
            return (
                f"📂 {data['category']}\n"
                f"Gasto: {self.format_currency_br(data['spent'])}\n"
                f"Limite: {self.format_currency_br(data['limit'])}\n"
                f"Saldo: {self.format_currency_br(data['remaining'])}\n"
                f"Status: {data['status']}",
                data,
            )

        if msg in {"/limites", "limites"}:
            return self.limits_message(), None

        if msg in {"/contas", "contas"}:
            return self.bills_message(days_ahead=self.alert_days()), None

        if msg.startswith("/pagar ") or msg.startswith("pagar "):
            bill_name = raw.split(" ", 1)[1]
            try:
                result = self.mark_bill_paid(bill_name)
            except ValueError as exc:
                return str(exc), None
            return f"💸 Conta marcada como paga.\n{result.message}", result.__dict__

        if msg.startswith("/gasto ") or msg.startswith("/receita "):
            entry_type = "receita" if msg.startswith("/receita ") else "gasto"
            body = raw.split(" ", 1)[1]
            launch_pattern = r"^([^\d]+?)\s+(\d+(?:[.,]\d{1,2})?)(?:\s+(.*))?$"
            match = re.match(launch_pattern, body, flags=re.IGNORECASE)
            if not match:
                return "Formato inválido. Exemplo: /gasto gasolina 120 shell", None
            category_raw, amount_raw, desc = match.groups()
            amount = float(amount_raw.replace(".", "").replace(",", ".")) if "," in amount_raw else float(amount_raw)
            result = self.add_launch(category=category_raw, amount=amount, description=desc or "", entry_type=entry_type, channel="telegram")
            return result.message, result.__dict__

        launch_pattern = r"^([^\d/]+?)\s+(\d+(?:[.,]\d{1,2})?)(?:\s+(.*))?$"
        match = re.match(launch_pattern, raw, flags=re.IGNORECASE)
        if match:
            category_raw, amount_raw, desc = match.groups()
            amount = float(amount_raw.replace(".", "").replace(",", ".")) if "," in amount_raw else float(amount_raw)
            result = self.add_launch(category=category_raw, amount=amount, description=desc or "", entry_type="gasto", channel="telegram")
            return result.message, result.__dict__

        return "Não entendi. Use /ajuda para ver os comandos.", None
